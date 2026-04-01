# distributorplatform/app/order/views.py
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from django.shortcuts import render, redirect, get_object_or_404
from django.http import Http404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction, IntegrityError
from django.urls import reverse
from django.utils import timezone
from decimal import Decimal
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

from django.core.paginator import Paginator
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Q, Prefetch, Max, Sum, F, DecimalField
from django.db.models.functions import Coalesce, TruncDate
from datetime import datetime
import json
import hashlib
import urllib.parse

from inventory.models import QuotationItem
from product.models import Product, Category
from .models import Order, OrderItem, Customer, CustomerAddress
from .forms import ManualOrderForm
from core.models import SiteSetting, PaymentOption

ORDER_EXPORT_HEADERS = [
    'Order ID', 'Order Date', 'Salesteam', 'Customer Name', 'Product Name',
    'Product Base Cost', 'Platform Price', 'Actual Received (Unit)',
    'Quantity', 'Line Revenue', 'Profit'
]
MONTH_FILL_BLUE = PatternFill(fill_type='solid', fgColor='E6F2FF')
MONTH_FILL_WHITE = PatternFill(fill_type='solid', fgColor='FFFFFF')
HEADER_FILL_GRAY = PatternFill(fill_type='solid', fgColor='F2F2F2')
HEADER_FONT_BOLD = Font(bold=True)


def _logical_order_date(order):
    return order.transaction_date or (order.created_at.date() if order.created_at else None)


def _excel_response_for_order_rows(filename, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Orders'
    ws.append(ORDER_EXPORT_HEADERS)
    for col in range(1, len(ORDER_EXPORT_HEADERS) + 1):
        header_cell = ws.cell(row=1, column=col)
        header_cell.fill = HEADER_FILL_GRAY
        header_cell.font = HEADER_FONT_BOLD

    prev_month = None
    use_blue = True
    for row in rows:
        month_key = row.get('month_key')
        if prev_month is not None and month_key != prev_month:
            use_blue = not use_blue
        fill = MONTH_FILL_BLUE if use_blue else MONTH_FILL_WHITE

        ws.append(row['values'])
        current_row = ws.max_row
        for col in range(1, len(ORDER_EXPORT_HEADERS) + 1):
            ws.cell(row=current_row, column=col).fill = fill

        prev_month = month_key

    # Auto-fit column widths based on max content length (with a small cap).
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            value = '' if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 48)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def salesperson_required(view_func):
    """
    Allow only superusers or users in sales-related groups
    (e.g. 'Salesperson' or 'salesteam').
    """
    @login_required
    def _wrapped(request, *args, **kwargs):
        if request.user.is_superuser:
            return view_func(request, *args, **kwargs)
        if request.user.user_groups.filter(
            Q(name__iexact='Salesperson') | Q(name__iexact='salesteam')
        ).exists():
            return view_func(request, *args, **kwargs)
        messages.error(request, 'You do not have permission to access Manual Order Entry.')
        return redirect(reverse('core:manage_dashboard'))
    return _wrapped


def agent_required(view_func):
    """
    Decorator to ensure the user is logged in.
    """
    @login_required
    def _wrapped_view(request, *args, **kwargs):
        return view_func(request, *args, **kwargs)
    return _wrapped_view


@staff_member_required
def export_customers_from_orders(request):
    """
    Export implicit customers derived from orders as CSV.
    Group by (customer_name, customer_phone, customer_email, shipping_address).
    """
    # Only staff should hit this view by decorator; keep logic simple.
    # Collect unique customer signatures
    seen = {}
    qs = Order.objects.all().order_by('created_at')
    for o in qs.iterator():
        name = (o.customer_name or '').strip()
        phone = (o.customer_phone or '').strip()
        # You don't currently store email on Order; keep placeholder for future
        email = ''
        address = (o.shipping_address or '').strip()
        if not (name or phone or address):
            continue
        key = (name.lower(), phone, email.lower(), address)
        if key not in seen:
            seen[key] = {
                'name': name,
                'phone': phone,
                'email': email,
                'address': address,
                'notes': 'Imported from orders',
            }

    # Stream CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="customers_from_orders.csv"'
    writer = csv.writer(response)
    writer.writerow(['name', 'phone', 'email', 'address', 'notes'])
    for row in seen.values():
        writer.writerow([row['name'], row['phone'], row['email'], row['address'], row['notes']])
    return response


@agent_required
def place_order_view(request):
    """
    Displays the main "Place Order" interface.
    Accessible by any logged-in user (Customer or Agent).
    """
    allowed_categories = Category.objects.filter(
        user_groups__users=request.user
    )

    products_query = Product.objects.filter(
        categories__in=allowed_categories
    ).select_related('featured_image').prefetch_related(
        Prefetch(
            'quotationitem_set',
            queryset=QuotationItem.objects.select_related('quotation')
                                          .prefetch_related('quotation__items')
                                          .order_by('-quotation__date_quoted'),
            to_attr='latest_quotation_items'
        ),
        'price_tiers',
    ).distinct().order_by('name')

    # --- 1. Determine Agent Status ---
    # User is an 'Agent' ONLY if they belong to a group with > 0% commission.
    max_commission_data = request.user.user_groups.aggregate(Max('commission_percentage'))
    agent_commission_percent = max_commission_data['commission_percentage__max'] or Decimal('0.00')
    is_agent = agent_commission_percent > 0

    product_list_json = []
    for p in products_query:
        selling_price = p.selling_price
        base_cost = p.base_cost

        if selling_price is not None:
            # Price tiers for tiered pricing (min_quantity -> price per unit)
            price_tiers = [{'min_quantity': t.min_quantity, 'price': float(t.price)} for t in p.price_tiers.all()]
            price_tiers.sort(key=lambda x: -x['min_quantity'])  # highest min_quantity first

            item_data = {
                'id': p.id,
                'name': p.name,
                'sku': p.sku or '-',
                'selling_price': float(selling_price),
                'price_tiers': price_tiers,
                'img_url': p.featured_image.image.url if p.featured_image else None,
                'commission': 0.00,  # Default for customers
            }

            if is_agent:
                # --- 2. Calculate Agent's Commission (per unit at base price) ---
                commission_value = Decimal('0.00')
                if p.profit_margin is not None:
                    profit_base = selling_price * (p.profit_margin / Decimal('100.00'))
                elif base_cost is not None:
                    profit_base = selling_price - base_cost
                else:
                    profit_base = Decimal('0.00')
                if profit_base > 0:
                    commission_value = profit_base * (agent_commission_percent / Decimal('100.00'))
                item_data['commission'] = float(commission_value)

            product_list_json.append(item_data)

    context = {
        'products_json': product_list_json,
        'is_agent': is_agent,
    }
    return render(request, 'order/place_order.html', context)


@login_required
@transaction.atomic
def api_submit_order(request):
    """
    API endpoint to receive the cart data and create an order.
    """
    if request.method != 'POST' or not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=400)

    try:
        data = json.loads(request.body)
        cart_items = data.get('cart', [])

        if not cart_items:
            return JsonResponse({'success': False, 'error': 'Your cart is empty.'}, status=400)

        # 1. Create the parent Order
        new_order = Order.objects.create(
            agent=request.user,
            status=Order.OrderStatus.PENDING
        )

        items_to_create = []
        product_ids = [item.get('id') for item in cart_items]

        # 2. Re-fetch products to get current prices (Security)
        products_in_cart = Product.objects.filter(id__in=product_ids).prefetch_related(
             Prefetch(
                'quotationitem_set',
                queryset=QuotationItem.objects.select_related('quotation')
                                              .prefetch_related('quotation__items')
                                              .order_by('-quotation__date_quoted'),
                to_attr='latest_quotation_items'
            ),
             'price_tiers',
        )

        product_map = {p.id: p for p in products_in_cart}

        # 3. Validate cart
        for item in cart_items:
            product_id = item.get('id')
            quantity = int(item.get('quantity', 0))

            if quantity <= 0: continue

            if product_id not in product_map:
                raise IntegrityError(f"Product {product_id} unavailable.")

            product = product_map[product_id]

            unit_price = product.get_price_for_quantity(quantity)
            if unit_price is None:
                raise IntegrityError(f"Product {product.name} has no selling price.")

            landed_cost = product.base_cost if product.base_cost is not None else Decimal('0.00')

            items_to_create.append(
                OrderItem(
                    order=new_order,
                    product=product,
                    quantity=quantity,
                    selling_price=unit_price,
                    landed_cost=landed_cost
                )
            )

        if not items_to_create:
            raise IntegrityError("No valid items were found in the cart.")

        for item in items_to_create:
            item.save()

        success_url = reverse('order:order_success', kwargs={'order_id': new_order.id})
        return JsonResponse({'success': True, 'redirect_url': success_url})

    except IntegrityError as e:
        transaction.set_rollback(True)
        return JsonResponse({'success': False, 'error': f'Order processing failed: {e}'}, status=400)
    except Exception as e:
        transaction.set_rollback(True)
        return JsonResponse({'success': False, 'error': f'An unexpected error occurred: {e}'}, status=500)


# --- Manual Order Entry (Salesperson) ---

@salesperson_required
def create_manual_order_view(request):
    """Manual order entry: form for sales channel + customer details; product selection with custom unit_price."""
    form = ManualOrderForm()
    context = {
        'form': form,
        'sales_channel_choices': Order.SalesChannel.choices,
    }
    return render(request, 'order/create_manual_order.html', context)


@salesperson_required
def api_manual_order_products(request):
    """GET: list products for manual order (id, name, sku, default selling_price, base_cost)."""
    if request.method != 'GET':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    allowed_categories = Category.objects.filter(user_groups__users=request.user)
    if request.user.is_superuser or not allowed_categories.exists():
        products_query = Product.objects.all()
    else:
        products_query = Product.objects.filter(categories__in=allowed_categories)
    products_query = products_query.distinct().order_by('name')
    product_list = []
    for p in products_query:
        product_list.append({
            'id': p.id,
            'name': p.name,
            'alias_name': p.alias_name or '',
            'display_name': p.order_display_name,
            'sku': p.sku or '-',
            'selling_price': str(p.selling_price) if p.selling_price is not None else None,
            'base_cost': str(p.base_cost) if p.base_cost is not None else '0.00',
        })
    return JsonResponse({'success': True, 'products': product_list})


@salesperson_required
def api_customer_search(request):
    """GET: search customers by name/phone/email for manual order picker. ?q=..."""
    if request.method != 'GET':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    q = (request.GET.get('q') or '').strip()[:100]
    customers = Customer.objects.all().order_by('name')
    if q:
        customers = customers.filter(
            Q(name__icontains=q) |
            Q(phone__icontains=q) |
            Q(email__icontains=q) |
            Q(address__icontains=q)
        )[:15]
    else:
        customers = customers[:20]
    items = [
        {
            'id': c.id,
            'name': c.name or '',
            'phone': c.phone or '',
            'email': c.email or '',
            'address': c.address or '',
        }
        for c in customers
    ]
    return JsonResponse({'success': True, 'customers': items})


@staff_member_required
def manage_customers_dashboard(request):
    """Manage Customers page: list, add, edit, delete central customer records. Uses manage tools sidebar."""
    return render(request, 'order/manage_customers.html', {
        'title': 'Manage Customers',
        'is_subpage': True,
    })


@staff_member_required
def api_customers_list(request):
    """GET: paginated list of customers for manage page. ?page=1&search=..."""
    if request.method != 'GET':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    search = (request.GET.get('search') or '').strip()[:100]
    qs = Customer.objects.all().order_by('name')
    if search:
        qs = qs.filter(
            Q(name__icontains=search) |
            Q(phone__icontains=search) |
            Q(email__icontains=search)
        )
    paginator = Paginator(qs, 25)
    page_num = request.GET.get('page', '1')
    try:
        page = paginator.page(int(page_num))
    except (ValueError, TypeError):
        page = paginator.page(1)
    items = [
        {
            'id': c.id,
            'name': c.name or '',
            'phone': c.phone or '',
            'email': c.email or '',
            'address': (c.address or '')[:80] + ('...' if c.address and len(c.address) > 80 else ''),
            'notes': (c.notes or '')[:80] + ('...' if c.notes and len(c.notes) > 80 else ''),
            'created_at': c.created_at.strftime('%Y-%m-%d %H:%M'),
            'order_count': c.orders.count(),
        }
        for c in page.object_list
    ]
    return JsonResponse({
        'success': True,
        'items': items,
        'page': page.number,
        'total_pages': paginator.num_pages,
        'total_count': paginator.count,
    })


@staff_member_required
@transaction.atomic
def api_customer_create(request):
    """POST: create a new customer. JSON: name, phone?, email?, address?, notes?"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
    name = (data.get('name') or '').strip()
    if not name:
        return JsonResponse({'success': False, 'error': 'Name is required.'}, status=400)
    customer = Customer.objects.create(
        name=name,
        phone=(data.get('phone') or '').strip() or None,
        email=(data.get('email') or '').strip() or None,
        address=(data.get('address') or '').strip() or None,
        notes=(data.get('notes') or '').strip() or None,
    )
    return JsonResponse({
        'success': True,
        'customer': {
            'id': customer.id,
            'name': customer.name,
            'phone': customer.phone or '',
            'email': customer.email or '',
            'address': customer.address or '',
            'notes': customer.notes or '',
        },
    })


@staff_member_required
@transaction.atomic
def api_customer_update(request, customer_id):
    """POST: update customer. JSON: name?, phone?, email?, address?, notes?"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    customer = get_object_or_404(Customer, id=customer_id)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
    if data.get('name') is not None:
        name = (data.get('name') or '').strip()
        if not name:
            return JsonResponse({'success': False, 'error': 'Name cannot be empty.'}, status=400)
        customer.name = name
    if 'phone' in data:
        customer.phone = (data.get('phone') or '').strip() or None
    if 'email' in data:
        customer.email = (data.get('email') or '').strip() or None
    if 'address' in data:
        customer.address = (data.get('address') or '').strip() or None
    if 'notes' in data:
        customer.notes = (data.get('notes') or '').strip() or None
    customer.save()
    return JsonResponse({
        'success': True,
        'customer': {
            'id': customer.id,
            'name': customer.name,
            'phone': customer.phone or '',
            'email': customer.email or '',
            'address': customer.address or '',
            'notes': customer.notes or '',
        },
    })


@staff_member_required
@transaction.atomic
def api_customer_delete(request, customer_id):
    """POST: delete a customer (only if no orders linked, or allow and set orders.customer=None)."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    customer = get_object_or_404(Customer, id=customer_id)
    order_count = customer.orders.count()
    # We allow delete; orders keep their snapshot (customer_name, etc.) via SET_NULL
    customer.delete()
    return JsonResponse({
        'success': True,
        'message': f'Customer deleted. {order_count} order(s) were unlinked.',
    })


@staff_member_required
def api_customer_orders(request, customer_id):
    """
    GET: list recent orders for a specific customer (for Manage Customers modal).
    Combines orders linked via FK and, as a fallback, snapshot name/phone.
    """
    if request.method != 'GET':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    customer = get_object_or_404(Customer, id=customer_id)

    # Only use FK-linked orders for now to avoid union quirks
    qs = (
        Order.objects
        .filter(customer_id=customer.id)
        .select_related('agent')
        .prefetch_related('items')
        .order_by('-created_at')[:50]
    )

    orders_data = []
    for o in qs:
        items = list(o.items.all())
        total = sum((item.total_price for item in items), Decimal('0.00'))
        orders_data.append({
            'id': str(o.id),
            'created_at': o.created_at.strftime('%Y-%m-%d %H:%M'),
            'transaction_date': o.transaction_date.isoformat() if o.transaction_date else '',
            'status': o.status,
            'status_display': o.get_status_display(),
            'sales_channel': o.sales_channel or '',
            'customer_name': o.customer_name or '',
            'customer_phone': o.customer_phone or '',
            'total': f'{total:.2f}',
        })

    return JsonResponse({
        'success': True,
        'customer': {
            'id': customer.id,
            'name': customer.name,
            'phone': customer.phone or '',
        },
        'orders': orders_data,
    })


@staff_member_required
def api_customer_detail(request, customer_id):
    """GET: single customer for edit form."""
    if request.method != 'GET':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    customer = get_object_or_404(Customer, id=customer_id)
    return JsonResponse({
        'success': True,
        'customer': {
            'id': customer.id,
            'name': customer.name,
            'phone': customer.phone or '',
            'email': customer.email or '',
            'address': customer.address or '',
            'notes': customer.notes or '',
        },
    })


@salesperson_required
@transaction.atomic
def api_submit_manual_order(request):
    """
    POST: create a manual order (no commission).
    Expects JSON: sales_channel, customer_name, customer_phone, shipping_address,
    items: [{ product_id, quantity, unit_price }]
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    items = data.get('items') or []
    if not items:
        return JsonResponse({'success': False, 'error': 'At least one order item is required.'}, status=400)

    product_ids = [int(i.get('product_id')) for i in items if i.get('product_id')]
    if not product_ids:
        return JsonResponse({'success': False, 'error': 'Invalid items.'}, status=400)

    products = Product.objects.filter(id__in=product_ids)
    product_map = {p.id: p for p in products}

    sales_channel = data.get('sales_channel') or Order.SalesChannel.OTHER
    if sales_channel not in dict(Order.SalesChannel.choices):
        sales_channel = Order.SalesChannel.OTHER

    transaction_date = None
    if data.get('transaction_date'):
        try:
            from datetime import datetime
            transaction_date = datetime.strptime(data['transaction_date'], '%Y-%m-%d').date()
        except (ValueError, TypeError):
            pass

    customer = None
    customer_name = (data.get('customer_name') or '').strip() or None
    customer_phone = (data.get('customer_phone') or '').strip() or None
    shipping_address = (data.get('shipping_address') or '').strip() or None
    customer_id = data.get('customer_id')
    if customer_id:
        try:
            customer = Customer.objects.get(id=int(customer_id))
            customer_name = customer.name
            customer_phone = customer.phone or customer_phone
            # Keep order's shipping_address payload as provided; don't force override with customer.address
        except (ValueError, TypeError, Customer.DoesNotExist):
            customer = None
    elif customer_name or customer_phone or shipping_address:
        customer = Customer.objects.create(
            name=customer_name or 'Unknown',
            phone=customer_phone,
            email=None,
            address=shipping_address,
            notes=None,
        )

    # Snapshot fields on order always use the latest values
    customer_name = customer_name or (customer.name if customer else None)
    customer_phone = customer_phone or (customer.phone if customer else None)

    # Persist new shipping address variant in address book (without overwriting customer.address)
    if customer and shipping_address:
        addr_text = shipping_address.strip()
        if addr_text:
            exists = CustomerAddress.objects.filter(customer=customer, address=addr_text).exists()
            if not exists:
                is_default = not CustomerAddress.objects.filter(customer=customer).exists()
                CustomerAddress.objects.create(
                    customer=customer,
                    label=None,
                    address=addr_text,
                    is_default=is_default,
                )

    new_order = Order.objects.create(
        agent=request.user,
        created_by=request.user,
        sales_channel=sales_channel,
        transaction_date=transaction_date,
        customer=customer,
        customer_name=customer_name,
        customer_phone=customer_phone,
        shipping_address=shipping_address,
        remarks=(data.get('remarks') or '').strip() or None,
        status=Order.OrderStatus.PENDING,
    )

    order_items_to_create = []
    for row in items:
        product_id = row.get('product_id')
        try:
            product_id = int(product_id)
        except (TypeError, ValueError):
            continue
        if product_id not in product_map:
            continue
        product = product_map[product_id]
        quantity = int(row.get('quantity') or 0)
        if quantity <= 0:
            continue
        try:
            actual_unit_price = Decimal(str(row.get('actual_unit_price') or row.get('unit_price') or 0))
        except Exception:
            actual_unit_price = product.selling_price or Decimal('0.00')
        if actual_unit_price < 0:
            actual_unit_price = Decimal('0.00')
        try:
            platform_price = Decimal(str(row.get('platform_price') or 0)) if row.get('platform_price') not in (None, '') else None
        except Exception:
            platform_price = None
        if platform_price is not None and platform_price < 0:
            platform_price = None
        landed_cost = product.base_cost if product.base_cost is not None else Decimal('0.00')

        order_items_to_create.append(
            OrderItem(
                order=new_order,
                product=product,
                quantity=quantity,
                selling_price=actual_unit_price,
                landed_cost=landed_cost,
                platform_price=platform_price,
                actual_unit_price=actual_unit_price,
            )
        )

    if not order_items_to_create:
        new_order.delete()
        return JsonResponse({'success': False, 'error': 'No valid items.'}, status=400)

    for oi in order_items_to_create:
        oi.save()

    success_url = reverse('order:order_success', kwargs={'order_id': new_order.id})
    # Prepare lightweight summary for inline success modal (manual order entry)
    order_items = new_order.items.select_related('product')
    subtotal = sum(item.total_price for item in order_items)
    site_settings = SiteSetting.load()
    cs_phone = site_settings.customer_service_whatsapp

    msg_lines = [
        "*New Manual Order Recorded!*",
        f"Order ID: #{new_order.id}",
        f"Agent: {request.user.username}",
        f"Date: {new_order.created_at.strftime('%Y-%m-%d %H:%M')}",
        "",
        "*Items:*",
    ]
    items_summary = []
    for item in order_items:
        disp = item.product.order_display_name
        msg_lines.append(f"- {disp} (x{item.quantity})")
        items_summary.append({
            'product_name': disp,
            'quantity': item.quantity,
            'unit_price': str(item.selling_price),
            'line_total': str(item.total_price),
        })
    msg_lines.append("")
    msg_lines.append(f"*Total Amount: RM {subtotal:.2f}*")
    msg_lines.append(f"Status: {new_order.get_status_display()}")

    full_message = "\n".join(msg_lines)
    whatsapp_url = None
    if cs_phone:
        encoded_message = urllib.parse.quote(full_message)
        whatsapp_url = f"https://wa.me/{cs_phone}?text={encoded_message}"

    return JsonResponse({
        'success': True,
        'redirect_url': success_url,
        'order_id': new_order.id,
        'order': {
            'id': new_order.id,
            'items': items_summary,
            'subtotal': str(subtotal),
            'whatsapp_url': whatsapp_url,
        },
    })


@salesperson_required
@transaction.atomic
def api_manual_order_detail(request, order_id):
    """
    GET: return details for a manual order that can be edited by the salesperson.
    Only pending orders created_by the current user are editable.
    """
    if request.method != 'GET':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    allowed_statuses = [
        Order.OrderStatus.PENDING,
        Order.OrderStatus.TO_PAY,
        Order.OrderStatus.TO_SHIP,
        Order.OrderStatus.TO_RECEIVE,
    ]
    order = get_object_or_404(
        Order.objects.prefetch_related('items__product'),
        id=order_id,
        created_by=request.user,
        status__in=allowed_statuses,
    )

    items_data = []
    for item in order.items.all():
        items_data.append({
            'product_id': item.product_id,
            'sku': item.product.sku or '',
            'name': item.product.name,
            'display_name': item.product.order_display_name,
            'quantity': item.quantity,
            'platform_price': str(item.platform_price) if item.platform_price is not None else '',
            'actual_unit_price': str(item.actual_unit_price or item.selling_price),
        })

    data = {
        'id': order.id,
        'status': order.status,
        'sales_channel': order.sales_channel or Order.SalesChannel.OTHER,
        'transaction_date': order.transaction_date.isoformat() if order.transaction_date else '',
        'customer_id': order.customer_id,
        'customer_name': order.customer_name or '',
        'customer_phone': order.customer_phone or '',
        'shipping_address': order.shipping_address or '',
        'items': items_data,
        'remarks': order.remarks or '',
    }
    return JsonResponse({'success': True, 'order': data})


@salesperson_required
@transaction.atomic
def api_update_manual_order(request, order_id):
    """
    POST: update a pending manual order created by the current salesperson.
    Replaces the existing items with the provided list.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    allowed_statuses = [
        Order.OrderStatus.PENDING,
        Order.OrderStatus.TO_PAY,
        Order.OrderStatus.TO_SHIP,
        Order.OrderStatus.TO_RECEIVE,
    ]
    order = get_object_or_404(
        Order,
        id=order_id,
        created_by=request.user,
        status__in=allowed_statuses,
    )

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    items = data.get('items') or []
    if not items:
        return JsonResponse({'success': False, 'error': 'At least one order item is required.'}, status=400)

    product_ids = [int(i.get('product_id')) for i in items if i.get('product_id')]
    if not product_ids:
        return JsonResponse({'success': False, 'error': 'Invalid items.'}, status=400)

    # Map payload by product_id for quick lookup
    payload_by_product = {}
    for row in items:
        pid = row.get('product_id')
        try:
            pid = int(pid)
        except (TypeError, ValueError):
            continue
        payload_by_product[pid] = row

    if order.status == Order.OrderStatus.PENDING:
        # Full edit allowed (existing behavior)
        products = Product.objects.filter(id__in=product_ids)
        product_map = {p.id: p for p in products}

        sales_channel = data.get('sales_channel') or Order.SalesChannel.OTHER
        if sales_channel not in dict(Order.SalesChannel.choices):
            sales_channel = Order.SalesChannel.OTHER

        transaction_date = None
        if data.get('transaction_date'):
            try:
                from datetime import datetime
                transaction_date = datetime.strptime(data['transaction_date'], '%Y-%m-%d').date()
            except (ValueError, TypeError):
                pass

        # Update order header and customer
        order.sales_channel = sales_channel
        order.transaction_date = transaction_date
        order.remarks = (data.get('remarks') or '').strip() or None
        customer = None
        customer_name = (data.get('customer_name') or '').strip() or None
        customer_phone = (data.get('customer_phone') or '').strip() or None
        shipping_address = (data.get('shipping_address') or '').strip() or None
        customer_id = data.get('customer_id')
        if customer_id:
            try:
                customer = Customer.objects.get(id=int(customer_id))
                customer_name = customer_name or customer.name
                customer_phone = customer_phone or customer.phone
            except (ValueError, TypeError, Customer.DoesNotExist):
                pass
        elif customer_name or customer_phone or shipping_address:
            customer = Customer.objects.create(
                name=customer_name or 'Unknown',
                phone=customer_phone,
                email=None,
                address=shipping_address,
                notes=None,
            )
        # snapshot fields from final values / customer
        customer_name = customer_name or (customer.name if customer else None)
        customer_phone = customer_phone or (customer.phone if customer else None)

        # maintain address book
        if customer and shipping_address:
            addr_text = shipping_address.strip()
            if addr_text:
                exists = CustomerAddress.objects.filter(customer=customer, address=addr_text).exists()
                if not exists:
                    is_default = not CustomerAddress.objects.filter(customer=customer).exists()
                    CustomerAddress.objects.create(
                        customer=customer,
                        label=None,
                        address=addr_text,
                        is_default=is_default,
                    )
        order.customer = customer
        order.customer_name = customer_name
        order.customer_phone = customer_phone
        order.shipping_address = shipping_address
        order.save()

        # Replace items
        order.items.all().delete()

        order_items_to_create = []
        for row in items:
            product_id = row.get('product_id')
            try:
                product_id = int(product_id)
            except (TypeError, ValueError):
                continue
            if product_id not in product_map:
                continue
            product = product_map[product_id]
            quantity = int(row.get('quantity') or 0)
            if quantity <= 0:
                continue
            try:
                actual_unit_price = Decimal(str(row.get('actual_unit_price') or row.get('unit_price') or 0))
            except Exception:
                actual_unit_price = product.selling_price or Decimal('0.00')
            if actual_unit_price < 0:
                actual_unit_price = Decimal('0.00')
            try:
                platform_price = Decimal(str(row.get('platform_price') or 0)) if row.get('platform_price') not in (None, '') else None
            except Exception:
                platform_price = None
            if platform_price is not None and platform_price < 0:
                platform_price = None
            landed_cost = product.base_cost if product.base_cost is not None else Decimal('0.00')

            order_items_to_create.append(
                OrderItem(
                    order=order,
                    product=product,
                    quantity=quantity,
                    selling_price=actual_unit_price,
                    landed_cost=landed_cost,
                    platform_price=platform_price,
                    actual_unit_price=actual_unit_price,
                )
            )

        if not order_items_to_create:
            return JsonResponse({'success': False, 'error': 'No valid items.'}, status=400)

        for oi in order_items_to_create:
            oi.save()
    else:
        # Restricted edit: only allow updating actual received amount
        for item in order.items.select_related('product').all():
            row = payload_by_product.get(item.product_id)
            if not row:
                continue
            try:
                new_actual = Decimal(str(row.get('actual_unit_price') or row.get('unit_price') or 0))
            except Exception:
                continue
            if new_actual < 0:
                new_actual = Decimal('0.00')
            item.actual_unit_price = new_actual
            # selling_price and quantity remain unchanged; profit recalculated in save()
            item.save()

    success_url = reverse('order:order_success', kwargs={'order_id': order.id})
    # Prepare lightweight summary for inline success modal (manual order entry update)
    order_items = order.items.select_related('product')
    subtotal = sum(item.total_price for item in order_items)
    site_settings = SiteSetting.load()
    cs_phone = site_settings.customer_service_whatsapp

    msg_lines = [
        "*Manual Order Updated!*",
        f"Order ID: #{order.id}",
        f"Agent: {request.user.username}",
        f"Date: {order.created_at.strftime('%Y-%m-%d %H:%M')}",
        "",
        "*Items:*",
    ]
    items_summary = []
    for item in order_items:
        disp = item.product.order_display_name
        msg_lines.append(f"- {disp} (x{item.quantity})")
        items_summary.append({
            'product_name': disp,
            'quantity': item.quantity,
            'unit_price': str(item.selling_price),
            'line_total': str(item.total_price),
        })
    msg_lines.append("")
    msg_lines.append(f"*Total Amount: RM {subtotal:.2f}*")
    msg_lines.append(f"Status: {order.get_status_display()}")

    full_message = "\n".join(msg_lines)
    whatsapp_url = None
    if cs_phone:
        encoded_message = urllib.parse.quote(full_message)
        whatsapp_url = f"https://wa.me/{cs_phone}?text={encoded_message}"

    return JsonResponse({
        'success': True,
        'redirect_url': success_url,
        'order_id': order.id,
        'order': {
            'id': order.id,
            'items': items_summary,
            'subtotal': str(subtotal),
            'whatsapp_url': whatsapp_url,
        },
    })


@login_required
def order_success_view(request, order_id):
    # Allow access if user is agent or created_by (salesperson)
    order = get_object_or_404(Order, id=order_id)
    if order.agent_id != request.user.id and (not order.created_by_id or order.created_by_id != request.user.id):
        raise Http404()

    # Calculate totals for the template
    order_items = order.items.select_related('product')
    subtotal = sum(item.total_price for item in order_items)

    # Prepare WhatsApp Message
    site_settings = SiteSetting.load()
    cs_phone = site_settings.customer_service_whatsapp

    msg_lines = [
        f"*New Order Placed!*",
        f"Order ID: #{order.id}",
        f"Agent: {request.user.username}",
        f"Date: {order.created_at.strftime('%Y-%m-%d %H:%M')}",
        "",
        "*Items:*",
    ]

    for item in order_items:
        msg_lines.append(f"- {item.product.order_display_name} (x{item.quantity})")

    msg_lines.append("")
    msg_lines.append(f"*Total Amount: RM {subtotal:.2f}*")
    msg_lines.append(f"Status: {order.get_status_display()}")

    full_message = "\n".join(msg_lines)
    encoded_message = urllib.parse.quote(full_message)
    whatsapp_url = f"https://wa.me/{cs_phone}?text={encoded_message}"

    context = {
        'order': order,
        'order_items': order_items,
        'subtotal': subtotal,
        'whatsapp_url': whatsapp_url,
    }
    return render(request, 'order/order_success.html', context)

def _user_can_manual_order(user):
    return user.is_superuser or user.user_groups.filter(
        Q(name__iexact='Salesperson') | Q(name__iexact='salesteam')
    ).exists()


@staff_member_required
def manage_orders_dashboard(request):
    """Renders the dedicated Order Management Dashboard with Statistics."""
    # We pass empty stats initially; they will be populated by the Alpine.js API call
    # which defaults to the current month on load.

    context = {
        'title': 'Manage Orders',
        'order_statuses': Order.OrderStatus.choices,
        'sales_channel_choices': Order.SalesChannel.choices,
        'show_manual_order_link': _user_can_manual_order(request.user),
        'stats': {
            'total_orders': 0,
            'pending_orders': 0,
            'completed_orders': 0,
            'revenue': 0,
        }
    }
    return render(request, 'order/manage_orders.html', context)


@staff_member_required
def api_manage_orders(request):
    """JSON API to fetch filtered and paginated orders."""
    page_number = request.GET.get('page', 1)
    limit = request.GET.get('limit', 20)
    sort_by = request.GET.get('sort_by', 'order_date')
    sort_dir = request.GET.get('sort_dir', 'desc')
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')

    # --- Date Filter Params (Period + Range) ---
    try:
        month = int(request.GET.get('month', 0))
        year = int(request.GET.get('year', 0))
    except ValueError:
        month = 0
        year = 0

    # Optional explicit date range (overrides month/year if provided)
    start_date = None
    end_date = None
    start_str = request.GET.get('start_date') or ''
    end_str = request.GET.get('end_date') or ''
    if start_str:
        try:
            start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
        except ValueError:
            start_date = None
    if end_str:
        try:
            end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
        except ValueError:
            end_date = None

    # --- Sorting ---
    # Order date (sort_by order_date or created_at): sort by transaction_date when set, else the local
    # calendar date of created_at — same rule as the Order date column.
    sort_field_map = {
        'id': 'id',
        'customer': 'customer_name',
        'status': 'status',
    }
    sort_prefix = '-' if sort_dir == 'desc' else ''

    base_qs = Order.objects.select_related('agent').prefetch_related('items')

    if sort_by == 'total_value':
        # Placeholder ordering; refined after filters via Python sort by line totals
        orders = base_qs.order_by('-created_at')
    elif sort_by in sort_field_map:
        orders = base_qs.order_by(f'{sort_prefix}{sort_field_map[sort_by]}')
    else:
        # order_date (default), created_at (legacy), or unknown key
        tz = timezone.get_current_timezone()
        orders = base_qs.annotate(
            _order_sort_date=Coalesce(
                'transaction_date',
                TruncDate('created_at', tzinfo=tz),
            )
        ).order_by(f'{sort_prefix}_order_sort_date', 'id')

    # 1. Apply Date Filter (if provided)
    # Date range takes precedence over month/year when supplied.
    if start_date:
        orders = orders.filter(
            Q(transaction_date__gte=start_date)
            |
            Q(transaction_date__isnull=True, created_at__date__gte=start_date)
        )
    if end_date:
        orders = orders.filter(
            Q(transaction_date__lte=end_date)
            |
            Q(transaction_date__isnull=True, created_at__date__lte=end_date)
        )
    if not (start_date or end_date) and month and year:
        # Filter by logical "order date": prefer transaction_date if set, else created_at
        orders = orders.filter(
            Q(transaction_date__year=year, transaction_date__month=month)
            |
            Q(transaction_date__isnull=True, created_at__year=year, created_at__month=month)
        )

    # 2. Apply Status Filter
    if status_filter:
        orders = orders.filter(status=status_filter)

    # 3. Apply Search Filter: supports Order ID (including "#123"), customer_name, and agent fields
    if search_query:
        normalized_id_term = search_query.lstrip('#').strip()
        search_q = (
            Q(id__icontains=search_query) |
            Q(customer_name__icontains=search_query) |
            Q(agent__username__icontains=search_query) |
            Q(agent__first_name__icontains=search_query) |
            Q(agent__last_name__icontains=search_query)
        )
        if normalized_id_term and normalized_id_term != search_query:
            search_q |= Q(id__icontains=normalized_id_term)
        if normalized_id_term.isdigit():
            search_q |= Q(id=int(normalized_id_term))
        orders = orders.filter(
            search_q
        )

    # --- Statistics Calculation (Scoped to current Month/Search filters) ---
    # We create a separate queryset for stats to respect Date/Search filters
    # but IGNORE Pagination and usually ignore status filter (to show full overview of the month).
    stats_qs = Order.objects.all()

    if start_date:
        stats_qs = stats_qs.filter(
            Q(transaction_date__gte=start_date)
            |
            Q(transaction_date__isnull=True, created_at__date__gte=start_date)
        )
    if end_date:
        stats_qs = stats_qs.filter(
            Q(transaction_date__lte=end_date)
            |
            Q(transaction_date__isnull=True, created_at__date__lte=end_date)
        )
    if not (start_date or end_date) and month and year:
        stats_qs = stats_qs.filter(
            Q(transaction_date__year=year, transaction_date__month=month)
            |
            Q(transaction_date__isnull=True, created_at__year=year, created_at__month=month)
        )

    if search_query:
        normalized_id_term = search_query.lstrip('#').strip()
        search_q = (
            Q(id__icontains=search_query) |
            Q(customer_name__icontains=search_query) |
            Q(agent__username__icontains=search_query) |
            Q(agent__first_name__icontains=search_query) |
            Q(agent__last_name__icontains=search_query)
        )
        if normalized_id_term and normalized_id_term != search_query:
            search_q |= Q(id__icontains=normalized_id_term)
        if normalized_id_term.isdigit():
            search_q |= Q(id=int(normalized_id_term))
        stats_qs = stats_qs.filter(
            search_q
        )

    total_orders = stats_qs.exclude(status=Order.OrderStatus.CANCELLED).count()
    pending_orders = stats_qs.exclude(status__in=[
        Order.OrderStatus.COMPLETED,
        Order.OrderStatus.CLOSED,
        Order.OrderStatus.CANCELLED
    ]).count()
    completed_orders = stats_qs.filter(status__in=[Order.OrderStatus.COMPLETED, Order.OrderStatus.CLOSED]).count()

    revenue = 0
    if request.user.is_superuser:
        # Calculate revenue for non-cancelled orders in this scope
        rev_qs = OrderItem.objects.filter(order__in=stats_qs).exclude(order__status=Order.OrderStatus.CANCELLED)
        rev_agg = rev_qs.aggregate(t=Sum(F('selling_price') * F('quantity'), output_field=DecimalField()))
        revenue = float(rev_agg['t'] or 0)

    stats = {
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        'completed_orders': completed_orders,
        'revenue': revenue,
    }

    # --- Pagination & Serialization ---
    orders = orders.select_related('created_by')

    # For sorts that depend on computed totals (e.g. total_value), apply Python-side sort
    if sort_by == 'total_value':
        orders_with_totals = []
        for o in orders:
            total_value = sum(item.selling_price * item.quantity for item in o.items.all())
            orders_with_totals.append((o, total_value))
        reverse = (sort_dir == 'desc')
        orders_with_totals.sort(key=lambda t: t[1], reverse=reverse)
        orders = [o for o, _ in orders_with_totals]

    paginator = Paginator(orders, limit)
    page_obj = paginator.get_page(page_number)

    data = []
    for order in page_obj:
        total_items = sum(item.quantity for item in order.items.all())
        total_value = sum(item.selling_price * item.quantity for item in order.items.all())
        gross_profit = sum(item.profit for item in order.items.all())
        customer_display = (order.customer_name and order.customer_name.strip()) or order.agent.username

        # Display date: prefer transaction_date (manual orders), else localized created_at date
        if order.transaction_date:
            display_date = order.transaction_date.strftime('%d/%m/%Y')
        else:
            local_dt = timezone.localtime(order.created_at)
            display_date = local_dt.strftime('%d/%m/%Y')

        data.append({
            'id': order.id,
            'created_at': display_date,
            'agent': order.agent.username,
            'customer': customer_display,
            'is_agent': False,
            'is_manual_order': bool(order.created_by_id),
            'sales_channel': order.sales_channel or '',
            'created_by_username': order.created_by.username if order.created_by_id else None,
            'status': order.get_status_display(),
            'status_code': order.status,
            'total_items': total_items,
            'total_value': float(total_value),
            'total_profit': float(gross_profit),
            'total_commission': float(order.total_commission),
        })

    return JsonResponse({
        'items': data,
        'pagination': {
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous(),
            'num_pages': paginator.num_pages,
            'current_page': page_obj.number,
        },
        'stats': stats
    })


@staff_member_required
def api_get_order_items(request, order_id):
    """JSON API to fetch items for a specific order (Accordion)."""
    order = get_object_or_404(Order, pk=order_id)
    items = order.items.select_related('product').all()

    data = []
    for item in items:
        data.append({
            'product_name': item.product.name,
            'summary_name': item.product.order_display_name,
            'sku': item.product.sku or '-',
            'quantity': item.quantity,
            'selling_price': float(item.selling_price),
            'total': float(item.selling_price * item.quantity),
            'profit': float(item.profit),
        })

    return JsonResponse({'items': data})


@staff_member_required
def api_manage_order_edit_details(request, order_id):
    """
    GET: order header fields for the manage-orders edit modal.
    POST: update customer snapshot, channel, transaction date, payment, remarks (staff).
    """
    order = get_object_or_404(Order, pk=order_id)

    if request.method == 'GET':
        data = {
            'id': order.id,
            'status': order.status,
            'status_display': order.get_status_display(),
            'is_manual_order': bool(order.created_by_id),
            'sales_channel': order.sales_channel or '',
            'transaction_date': order.transaction_date.isoformat() if order.transaction_date else '',
            'customer_name': order.customer_name or '',
            'customer_phone': order.customer_phone or '',
            'shipping_address': order.shipping_address or '',
            'remarks': order.remarks or '',
            'payment_method': order.payment_method or '',
        }
        return JsonResponse({'success': True, 'order': data})

    if request.method == 'POST':
        try:
            payload = json.loads(request.body or '{}')
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

        sc = payload.get('sales_channel', order.sales_channel)
        if sc not in dict(Order.SalesChannel.choices):
            sc = Order.SalesChannel.OTHER

        td_raw = payload.get('transaction_date')
        if td_raw in (None, '', False):
            new_td = None
        else:
            try:
                new_td = datetime.strptime(str(td_raw)[:10], '%Y-%m-%d').date()
            except (ValueError, TypeError):
                new_td = order.transaction_date

        order.sales_channel = sc
        order.transaction_date = new_td
        order.customer_name = (payload.get('customer_name') or '').strip() or None
        order.customer_phone = (payload.get('customer_phone') or '').strip() or None
        order.shipping_address = (payload.get('shipping_address') or '').strip() or None
        order.remarks = (payload.get('remarks') or '').strip() or None
        order.payment_method = (payload.get('payment_method') or '').strip() or None
        order.save()

        return JsonResponse({'success': True, 'message': 'Order updated'})

    return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)


@staff_member_required
@transaction.atomic
def api_update_order_status(request, order_id):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=400)

    try:
        data = json.loads(request.body)
        new_status = data.get('status')

        if new_status not in Order.OrderStatus.values:
             return JsonResponse({'success': False, 'error': 'Invalid status'}, status=400)

        order = get_object_or_404(Order, pk=order_id)
        order.status = new_status
        order.save()

        return JsonResponse({'success': True, 'message': 'Status updated'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@staff_member_required
@transaction.atomic
def api_bulk_update_order_status(request):
    """
    POST: bulk update status for multiple orders.
    Body: { "order_ids": [...], "status": "TO_SHIP" }
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=400)

    try:
        data = json.loads(request.body or '{}')
        order_ids = data.get('order_ids') or []
        new_status = data.get('status')

        if not order_ids or not isinstance(order_ids, list):
            return JsonResponse({'success': False, 'error': 'No order IDs provided'}, status=400)
        if new_status not in Order.OrderStatus.values:
            return JsonResponse({'success': False, 'error': 'Invalid status'}, status=400)

        updated = Order.objects.filter(id__in=order_ids).update(status=new_status)
        return JsonResponse({'success': True, 'message': f'Status updated for {updated} order(s).'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@staff_member_required
@require_http_methods(['POST'])
def export_selected_orders(request):
    """
    Export selected orders as an itemized CSV (one row per OrderItem).
    Accepts JSON body: { "order_ids": ["id1", "id2", ...] }.
    """
    try:
        data = json.loads(request.body)
        order_ids = data.get('order_ids') or []
    except (json.JSONDecodeError, TypeError):
        return JsonResponse({'success': False, 'error': 'Invalid JSON or missing order_ids'}, status=400)

    if not order_ids or not isinstance(order_ids, list):
        return JsonResponse({'success': False, 'error': 'No order IDs provided'}, status=400)

    # Limit to avoid huge exports
    order_ids = list(set(str(oid).strip() for oid in order_ids if oid))[:500]

    orders = list(
        Order.objects.filter(id__in=order_ids)
        .select_related('agent', 'created_by')
        .prefetch_related('items__product')
    )
    orders.sort(key=lambda o: (_logical_order_date(o) or datetime.min.date(), o.created_at))

    rows = []
    for order in orders:
        salesteam_username = order.created_by.username if order.created_by_id else ''
        customer_name = (
            (order.customer_name and order.customer_name.strip())
            or (order.agent.get_full_name() and order.agent.get_full_name().strip())
            or order.agent.username
        )
        order_date_obj = _logical_order_date(order)
        order_date = order_date_obj.strftime('%Y-%m-%d') if order_date_obj else ''
        month_key = order_date_obj.strftime('%Y-%m') if order_date_obj else ''

        for item in order.items.all():
            product = item.product
            base_cost = product.saved_base_cost if (product.saved_base_cost is not None) else item.landed_cost
            platform_price = item.platform_price if item.platform_price is not None else ''
            actual_received = item.actual_unit_price if item.actual_unit_price is not None else item.selling_price
            line_revenue = actual_received * item.quantity
            rows.append({
                'month_key': month_key,
                'values': [
                    order.id,
                    order_date,
                    salesteam_username,
                    customer_name or '',
                    product.name or '',
                    float(base_cost) if base_cost is not None else '',
                    float(platform_price) if platform_price != '' else '',
                    float(actual_received),
                    item.quantity,
                    float(line_revenue),
                    float(item.profit),
                ],
            })

    return _excel_response_for_order_rows('orders_export.xlsx', rows)


@staff_member_required
def export_orders_range(request):
    """
    Export all order items whose logical order date (transaction_date or created_at)
    falls within the given date range.
    GET params:
        start_date, end_date in YYYY-MM-DD format (at least one required).
    """
    start_str = request.GET.get('start_date') or ''
    end_str = request.GET.get('end_date') or ''

    if not start_str and not end_str:
        return HttpResponse("start_date or end_date is required.", status=400)

    start_date = None
    end_date = None
    try:
        if start_str:
            start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
        if end_str:
            end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse("Invalid date format. Use YYYY-MM-DD.", status=400)

    orders = Order.objects.select_related('agent', 'created_by').prefetch_related('items__product')

    if start_date:
        orders = orders.filter(
            Q(transaction_date__gte=start_date)
            |
            Q(transaction_date__isnull=True, created_at__date__gte=start_date)
        )
    if end_date:
        orders = orders.filter(
            Q(transaction_date__lte=end_date)
            |
            Q(transaction_date__isnull=True, created_at__date__lte=end_date)
        )

    # Sort by logical order date oldest -> newest
    # (transaction_date when present, otherwise created_at)
    orders = sorted(
        list(orders),
        key=lambda o: (o.transaction_date or (o.created_at.date() if o.created_at else datetime.min.date()), o.created_at)
    )

    filename_parts = ["orders_range"]
    if start_date:
        filename_parts.append(start_date.strftime('%Y%m%d'))
    if end_date:
        filename_parts.append(end_date.strftime('%Y%m%d'))
    filename = f'{"_".join(filename_parts)}.xlsx'

    rows = []
    for order in orders:
        salesteam_username = order.created_by.username if order.created_by_id else ''
        customer_name = (
            (order.customer_name and order.customer_name.strip())
            or (order.agent.get_full_name() and order.agent.get_full_name().strip())
            or order.agent.username
        )
        order_date_obj = _logical_order_date(order)
        order_date_str = order_date_obj.strftime('%Y-%m-%d') if order_date_obj else ''
        month_key = order_date_obj.strftime('%Y-%m') if order_date_obj else ''

        for item in order.items.all():
            product = item.product
            base_cost = getattr(product, 'saved_base_cost', None)
            if base_cost is None:
                base_cost = item.landed_cost
            platform_price = item.platform_price if item.platform_price is not None else ''
            actual_received = item.actual_unit_price if item.actual_unit_price is not None else item.selling_price
            line_revenue = actual_received * item.quantity
            rows.append({
                'month_key': month_key,
                'values': [
                    order.id,
                    order_date_str,
                    salesteam_username,
                    customer_name or '',
                    product.name or '',
                    float(base_cost) if base_cost is not None else '',
                    float(platform_price) if platform_price != '' else '',
                    float(actual_received),
                    item.quantity,
                    float(line_revenue),
                    float(item.profit),
                ],
            })

    return _excel_response_for_order_rows(filename, rows)


@login_required
def api_prepare_checkout(request):
    """
    Receives cart JSON, validates items/prices, and stores the checkout payload in Session.
    """
    if request.method != 'POST' or not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Invalid request.'}, status=400)

    try:
        data = json.loads(request.body)
        cart_items = data.get('cart', [])

        if not cart_items:
            return JsonResponse({'success': False, 'error': 'Cart is empty.'}, status=400)

        product_ids = [item.get('id') for item in cart_items]
        products = Product.objects.filter(id__in=product_ids).prefetch_related('price_tiers')
        product_map = {p.id: p for p in products}

        checkout_cart = []

        max_commission_data = request.user.user_groups.aggregate(Max('commission_percentage'))
        agent_commission_percent = max_commission_data['commission_percentage__max'] or Decimal('0.00')
        is_agent = agent_commission_percent > 0

        for item in cart_items:
            p_id = item.get('id')
            quantity = int(item.get('quantity', 0))

            if quantity <= 0: continue
            if p_id not in product_map: continue

            product = product_map[p_id]
            selling_price = product.get_price_for_quantity(quantity)
            selling_price = selling_price if selling_price is not None else Decimal('0.00')
            base_cost = product.base_cost if product.base_cost is not None else Decimal('0.00')

            estimated_commission = Decimal('0.00')
            if is_agent:
                if product.profit_margin is not None:
                    profit_base = selling_price * (product.profit_margin / Decimal('100.00'))
                elif base_cost is not None:
                    profit_base = selling_price - base_cost
                else:
                    profit_base = Decimal('0.00')

                if profit_base > 0:
                    estimated_commission = (profit_base * (agent_commission_percent / Decimal('100.00'))) * quantity

            checkout_cart.append({
                'product_id': product.id,
                'name': product.name,
                'sku': product.sku or '-',
                'quantity': quantity,
                'selling_price': str(selling_price),
                'base_cost': str(base_cost),
                'total_price': str(selling_price * quantity),
                'total_commission': str(estimated_commission),
                'img_url': product.featured_image.image.url if product.featured_image else None
            })

        if not checkout_cart:
            return JsonResponse({'success': False, 'error': 'No valid products found.'}, status=400)

        request.session['checkout_data'] = checkout_cart
        return JsonResponse({'success': True, 'redirect_url': reverse('order:checkout_confirmation')})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def checkout_confirmation_view(request):
    """
    Displays the Checkout Confirmation page using data from Session.
    """
    checkout_data = request.session.get('checkout_data')

    if not checkout_data:
        messages.error(request, "Your checkout session has expired. Please try again.")
        return redirect('order:place_order')

    total_amount = Decimal('0.00')
    total_commission = Decimal('0.00')
    formatted_items = []

    for item in checkout_data:
        t_price = Decimal(item['total_price'])
        t_comm = Decimal(item.get('total_commission', item.get('total_profit', '0.00')))

        total_amount += t_price
        total_commission += t_comm

        formatted_items.append({
            **item,
            'selling_price': Decimal(item['selling_price']),
            'total_price': t_price,
            'total_commission': t_comm
        })

    is_agent = request.user.user_groups.filter(commission_percentage__gt=0).exists()

    context = {
        'items': formatted_items,
        'total_amount': total_amount,
        'total_commission': total_commission,
        'is_agent': is_agent
    }
    return render(request, 'order/checkout.html', context)


@login_required
@transaction.atomic
def api_confirm_checkout(request):
    """
    Finalizes the order using the data stored in Session.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request.'}, status=400)

    checkout_data = request.session.get('checkout_data')
    if not checkout_data:
        return JsonResponse({'success': False, 'error': 'Session expired.'}, status=400)

    try:
        order = Order.objects.create(
            agent=request.user,
            status=Order.OrderStatus.PENDING
        )

        items_to_create = []

        for item in checkout_data:
            qty = int(item['quantity'])
            sp = Decimal(item['selling_price'])
            lc = Decimal(item['base_cost'])

            db_profit = (sp - lc) * qty

            items_to_create.append(OrderItem(
                order=order,
                product_id=item['product_id'],
                quantity=qty,
                selling_price=sp,
                landed_cost=lc,
                profit=db_profit
            ))

        for item in items_to_create:
            item.save()

        del request.session['checkout_data']

        return JsonResponse({
            'success': True,
            'redirect_url': reverse('order:order_success', kwargs={'order_id': order.id})
        })

    except Exception as e:
        transaction.set_rollback(True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def payment_view(request, order_id):
    order = get_object_or_404(Order, id=order_id, agent=request.user)

    if order.status != Order.OrderStatus.TO_PAY:
        messages.error(request, "This order is not eligible for payment.")
        return redirect('user:profile')

    if request.method == 'POST':
        method_id = request.POST.get('payment_method_id')
        try:
            payment_option = PaymentOption.objects.get(id=method_id, is_active=True)
        except (PaymentOption.DoesNotExist, ValueError):
            messages.error(request, "Invalid payment method selected.")
            return redirect('user:profile')

        if payment_option.option_type == 'GATEWAY':
            amount_rm = sum(item.selling_price * item.quantity for item in order.items.all())
            if amount_rm <= 0:
                messages.error(request, "Invalid order amount.")
                return redirect('user:profile')

            config = SiteSetting.objects.first()
            if not config or not config.payment_enabled:
                messages.error(request, "Online payments are currently disabled in settings.")
                return redirect('user:profile')

            if config.payment_provider == 'SENANGPAY':
                return _initiate_order_senangpay(request, order, amount_rm, config)
            else:
                messages.error(request, f"Provider {config.payment_provider} not supported.")
                return redirect('user:profile')

        elif payment_option.option_type == 'COD':
            order.payment_method = payment_option.name
            order.status = Order.OrderStatus.TO_SHIP
            order.save()
            messages.success(request, f"Order confirmed! Payment method: {payment_option.name}. Please pay upon delivery.")
            return redirect('user:profile')

        elif payment_option.option_type == 'MANUAL':
            order.payment_method = payment_option.name
            order.status = Order.OrderStatus.TO_SHIP
            order.save()
            messages.success(request, f"Order submitted. Please follow the instructions for {payment_option.name}.")
            return redirect('user:profile')

    messages.error(request, "Invalid request method.")
    return redirect('user:profile')


def _initiate_order_senangpay(request, order, amount_rm, config):
    amount_str = f"{amount_rm:.2f}"
    merchant_id = str(config.payment_category_code).strip()
    secret_key = str(config.payment_api_key).strip()

    detail = f"Order_{order.id}"
    order_id_ref = str(order.id)

    data_to_hash = f"{secret_key}{detail}{amount_str}{order_id_ref}"
    hashed_value = hashlib.md5(data_to_hash.encode()).hexdigest()

    name = order.agent.username
    email = order.agent.email
    phone = str(order.agent.phone_number).replace('+', '') if order.agent.phone_number else ''

    params = {
        'detail': detail,
        'amount': amount_str,
        'order_id': order_id_ref,
        'name': name,
        'email': email,
        'phone': phone,
        'hash': hashed_value,
    }

    query_string = urllib.parse.urlencode(params)
    base_url = config.payment_gateway_url.strip().rstrip('/')
    payment_url = f"{base_url}/{merchant_id}?{query_string}"

    return redirect(payment_url)


def payment_callback_view(request):
    status_id = request.GET.get('status_id')
    order_id = request.GET.get('order_id')
    msg = request.GET.get('msg')
    transaction_id = request.GET.get('transaction_id')
    received_hash = request.GET.get('hash')

    config = SiteSetting.objects.first()
    secret_key = str(config.payment_api_key).strip()

    data_to_verify = f"{secret_key}{status_id}{order_id}{transaction_id}{msg}"
    expected_hash = hashlib.md5(data_to_verify.encode()).hexdigest()

    if received_hash != expected_hash:
        messages.error(request, "Security verification failed (Invalid Hash).")
        return redirect('user:profile')

    if status_id == '1':
        try:
            order = Order.objects.get(id=order_id)
            if order.status == Order.OrderStatus.TO_PAY:
                order.status = Order.OrderStatus.TO_SHIP
                order.save()
                messages.success(request, f"Payment successful! Order #{order.id} is now being processed.")
            else:
                messages.info(request, f"Order #{order.id} payment status check completed.")
        except Order.DoesNotExist:
            messages.error(request, "Order not found during payment callback.")
    else:
        messages.error(request, f"Payment failed: {msg}")

    return redirect('user:profile')


@csrf_exempt
def payment_webhook_view(request):
    if request.method == 'POST':
        data = request.POST
        status_id = data.get('status_id')
        order_id = data.get('order_id')
        msg = data.get('msg')
        transaction_id = data.get('transaction_id')
        received_hash = data.get('hash')

        config = SiteSetting.objects.first()
        secret_key = str(config.payment_api_key).strip()
        data_to_verify = f"{secret_key}{status_id}{order_id}{transaction_id}{msg}"
        expected_hash = hashlib.md5(data_to_verify.encode()).hexdigest()

        if received_hash != expected_hash:
            return HttpResponse("Invalid Hash", status=403)

        try:
            order = Order.objects.get(id=order_id)
            if status_id == '1' and order.status == Order.OrderStatus.TO_PAY:
                order.status = Order.OrderStatus.TO_SHIP
                order.save()
            return HttpResponse("OK")
        except Order.DoesNotExist:
            return HttpResponse("Order Not Found", status=404)

    return HttpResponse("Invalid Method", status=405)

@staff_member_required
def export_order_statement(request):
    try:
        month = int(request.GET.get('month', timezone.now().month))
        year = int(request.GET.get('year', timezone.now().year))
        status = request.GET.get('status', '')

        orders = Order.objects.filter(
            created_at__year=year,
            created_at__month=month
        ).select_related('agent', 'created_by').prefetch_related('items__product').order_by('created_at')

        if status:
            orders = orders.filter(status=status)

        rows = []

        for order in orders:
            items = list(order.items.all())
            salesteam_username = order.created_by.username if order.created_by_id else ''
            customer_name = (
                (order.customer_name and order.customer_name.strip())
                or (order.agent.get_full_name() and order.agent.get_full_name().strip())
                or order.agent.username
            )
            order_date = order.transaction_date or (order.created_at.date() if order.created_at else None)
            order_date_str = order_date.strftime('%Y-%m-%d') if hasattr(order_date, 'strftime') else ''
            month_key = order_date.strftime('%Y-%m') if hasattr(order_date, 'strftime') else ''

            for item in items:
                product = item.product
                base_cost = getattr(product, 'saved_base_cost', None)
                if base_cost is None:
                    base_cost = item.landed_cost
                platform_price = item.platform_price if item.platform_price is not None else ''
                actual_received = item.actual_unit_price if item.actual_unit_price is not None else item.selling_price
                line_revenue = actual_received * item.quantity
                rows.append({
                    'month_key': month_key,
                    'values': [
                        order.id,
                        order_date_str,
                        salesteam_username,
                        customer_name or '',
                        product.name or '',
                        float(base_cost) if base_cost is not None else '',
                        float(platform_price) if platform_price != '' else '',
                        float(actual_received),
                        item.quantity,
                        float(line_revenue),
                        float(item.profit),
                    ],
                })

        filename = f"order_statement_{year}_{month:02d}.xlsx"
        return _excel_response_for_order_rows(filename, rows)

    except Exception as e:
        return HttpResponse(f"Error exporting Excel: {str(e)}", status=500)
