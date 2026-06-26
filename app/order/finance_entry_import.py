"""Parse and generate Excel templates for order finance entry bulk uploads."""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook


def _normalize_header(cell: Any) -> str:
    if cell is None:
        return ''
    return ' '.join(str(cell).replace('\xa0', ' ').strip().lower().split())


def _parse_decimal(val: Any) -> Decimal | None:
    if val is None or (isinstance(val, str) and not val.strip()):
        return None
    if isinstance(val, bool):
        return None
    if isinstance(val, Decimal):
        return val
    if isinstance(val, int):
        return Decimal(val)
    if isinstance(val, float):
        return Decimal(str(val))
    s = str(val).strip().replace('\xa0', '').replace(' ', '')
    if not s:
        return None
    if ',' in s and '.' not in s:
        s = s.replace(',', '.')
    elif ',' in s and '.' in s and s.rfind(',') > s.rfind('.'):
        s = s.replace('.', '').replace(',', '.')
    else:
        s = s.replace(',', '')
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _parse_date(val: Any) -> date | None:
    if val is None or (isinstance(val, str) and not val.strip()):
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _row_is_empty(row: tuple | list) -> bool:
    return all(cell is None or str(cell).strip() == '' for cell in row)


def _build_header_map(row: tuple | list, aliases: dict[str, str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(row):
        key = aliases.get(_normalize_header(cell))
        if key and key not in mapping:
            mapping[key] = idx
    return mapping


def _cell(row: tuple | list, idx: int | None) -> Any:
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _cell_str(row: tuple | list, idx: int | None) -> str:
    val = _cell(row, idx)
    if val is None:
        return ''
    return str(val).strip()


def _workbook_bytes(headers: list[str], example: list[Any] | None = None) -> bytes:
    rows = [example] if example else []
    return _workbook_bytes_from_rows(headers, rows)


def _workbook_bytes_from_rows(headers: list[str], data_rows: list[list[Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Upload'
    ws.append(headers)
    for row in data_rows:
        ws.append(row)
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 40)
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _load_rows(file_bytes: bytes) -> list[tuple]:
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    return [tuple(row) for row in ws.iter_rows(values_only=True)]


def _parse_rows(
    file_bytes: bytes,
    *,
    required_fields: set[str],
    aliases: dict[str, str],
    row_parser,
) -> tuple[list[dict], list[str]]:
    rows = _load_rows(file_bytes)
    header_map: dict[str, int] | None = None
    header_row_num = 0
    parsed: list[dict] = []
    errors: list[str] = []

    for row_num, row in enumerate(rows, start=1):
        if _row_is_empty(row):
            continue
        if header_map is None:
            header_map = _build_header_map(row, aliases)
            missing = required_fields - set(header_map)
            if missing:
                pretty = ', '.join(sorted(missing))
                return [], [f'Missing required column(s): {pretty}']
            header_row_num = row_num
            continue
        try:
            item = row_parser(row, header_map, row_num)
            if item is not None:
                parsed.append(item)
        except ValueError as exc:
            errors.append(str(exc))

    if header_map is None:
        return [], ['Worksheet is empty or has no header row.']
    if not parsed and not errors:
        return [], ['No data rows found below the header row.']
    return parsed, errors


CASH_BANK_ALIASES = {
    'transaction id': 'transaction_id',
    'transaction_id': 'transaction_id',
    'type': 'payment_type',
    'payment type': 'payment_type',
    'received from': 'received_from',
    'from': 'received_from',
    'payer': 'received_from',
    'collected by': 'collected_by',
    'agent': 'collected_by',
    'transaction date': 'transaction_date',
    'date': 'transaction_date',
    'amount': 'amount',
}

CASH_BANK_EXPORT_HEADERS = [
    'Transaction ID',
    'Type',
    'Received From',
    'Collected By',
    'Transaction Date',
    'Amount',
]

CASH_BANK_TYPE_EXPORT_LABELS = {
    'CASH': 'Cash received',
    'BANK': 'Bank transfer',
    'LOAN': 'Loan repayment',
}

CASH_BANK_TYPE_MAP = {
    'CASH': 'CASH',
    'CASH RECEIVED': 'CASH',
    'BANK': 'BANK',
    'BANK TRANSFER': 'BANK',
    'LOAN': 'LOAN',
    'LOAN REPAYMENT': 'LOAN',
}


def _parse_cash_bank_row(row, header_map, row_num):
    transaction_id = _cell_str(row, header_map.get('transaction_id')).upper()
    collected_by = _cell_str(row, header_map.get('collected_by'))

    if transaction_id:
        if not collected_by:
            raise ValueError(f'Row {row_num}: collected by is required when transaction ID is set.')
        return {
            'transaction_id': transaction_id,
            'collected_by': collected_by,
            'is_update': True,
        }

    raw_type = _cell_str(row, header_map.get('payment_type')).upper()
    payment_type = CASH_BANK_TYPE_MAP.get(raw_type)
    if not payment_type:
        raise ValueError(f'Row {row_num}: type must be CASH, BANK, or LOAN.')

    received_from = _cell_str(row, header_map.get('received_from'))
    if not received_from:
        raise ValueError(f'Row {row_num}: received from is required.')

    if not collected_by:
        raise ValueError(f'Row {row_num}: collected by is required.')

    tx_date = _parse_date(_cell(row, header_map.get('transaction_date')))
    if not tx_date:
        raise ValueError(f'Row {row_num}: transaction date is required (YYYY-MM-DD).')

    amount = _parse_decimal(_cell(row, header_map.get('amount')))
    if amount is None or amount <= 0:
        raise ValueError(f'Row {row_num}: amount must be greater than zero.')

    return {
        'payment_type': payment_type,
        'received_from': received_from,
        'collected_by': collected_by,
        'transaction_date': tx_date,
        'amount': amount,
        'is_update': False,
    }


def parse_cash_bank_receipt_upload(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    return _parse_rows(
        file_bytes,
        required_fields={'payment_type', 'received_from', 'collected_by', 'transaction_date', 'amount'},
        aliases=CASH_BANK_ALIASES,
        row_parser=_parse_cash_bank_row,
    )


def cash_bank_receipt_template_bytes() -> bytes:
    return _workbook_bytes(
        CASH_BANK_EXPORT_HEADERS,
        ['', 'CASH', 'John Doe', 'agent1', '2026-06-26', 1000.00],
    )


def cash_received_transactions_export_bytes(rows: list[list[Any]]) -> bytes:
    return _workbook_bytes_from_rows(CASH_BANK_EXPORT_HEADERS, rows)


COMMISSION_ALIASES = {
    'paid to': 'paid_to',
    'agent': 'paid_to',
    'payment date': 'payment_date',
    'date': 'payment_date',
    'amount': 'amount',
    'notes': 'notes',
    'reference': 'notes',
}


def _parse_commission_row(row, header_map, row_num):
    paid_to = _cell_str(row, header_map.get('paid_to'))
    if not paid_to:
        raise ValueError(f'Row {row_num}: paid to is required.')

    payment_date = _parse_date(_cell(row, header_map.get('payment_date')))
    if not payment_date:
        raise ValueError(f'Row {row_num}: payment date is required (YYYY-MM-DD).')

    amount = _parse_decimal(_cell(row, header_map.get('amount')))
    if amount is None or amount <= 0:
        raise ValueError(f'Row {row_num}: amount must be greater than zero.')

    notes = _cell_str(row, header_map.get('notes'))
    return {
        'paid_to': paid_to,
        'payment_date': payment_date,
        'amount': amount,
        'notes': notes,
    }


def parse_commission_payment_upload(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    return _parse_rows(
        file_bytes,
        required_fields={'paid_to', 'payment_date', 'amount'},
        aliases=COMMISSION_ALIASES,
        row_parser=_parse_commission_row,
    )


def commission_payment_template_bytes() -> bytes:
    return _workbook_bytes(
        ['Paid To', 'Payment Date', 'Amount', 'Notes'],
        ['Agent Name', '2026-06-26', 500.00, 'June commission'],
    )


REVENUE_ADJUSTMENT_ALIASES = {
    'type': 'adjustment_type',
    'adjustment type': 'adjustment_type',
    'reference': 'reference',
    'description': 'reference',
    'transaction date': 'transaction_date',
    'date': 'transaction_date',
    'amount': 'amount',
}

REVENUE_TYPE_MAP = {
    'COMMISSION_RELEASED': 'COMMISSION_RELEASED',
    'COMMISSION RELEASED': 'COMMISSION_RELEASED',
    'LOAN_INTEREST': 'LOAN_INTEREST',
    'LOAN INTEREST': 'LOAN_INTEREST',
    'INTEREST OF LOAN': 'LOAN_INTEREST',
}


def _parse_revenue_adjustment_row(row, header_map, row_num):
    raw_type = _cell_str(row, header_map.get('adjustment_type')).upper()
    adjustment_type = REVENUE_TYPE_MAP.get(raw_type)
    if not adjustment_type:
        raise ValueError(
            f'Row {row_num}: type must be COMMISSION_RELEASED or LOAN_INTEREST.'
        )

    reference = _cell_str(row, header_map.get('reference'))
    if not reference:
        raise ValueError(f'Row {row_num}: reference is required.')

    tx_date = _parse_date(_cell(row, header_map.get('transaction_date')))
    if not tx_date:
        raise ValueError(f'Row {row_num}: transaction date is required (YYYY-MM-DD).')

    amount = _parse_decimal(_cell(row, header_map.get('amount')))
    if amount is None or amount <= 0:
        raise ValueError(f'Row {row_num}: amount must be greater than zero.')

    return {
        'adjustment_type': adjustment_type,
        'reference': reference,
        'transaction_date': tx_date,
        'amount': amount,
    }


def parse_revenue_adjustment_upload(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    return _parse_rows(
        file_bytes,
        required_fields={'adjustment_type', 'reference', 'transaction_date', 'amount'},
        aliases=REVENUE_ADJUSTMENT_ALIASES,
        row_parser=_parse_revenue_adjustment_row,
    )


def revenue_adjustment_template_bytes() -> bytes:
    return _workbook_bytes(
        ['Type', 'Reference', 'Transaction Date', 'Amount'],
        ['COMMISSION_RELEASED', 'Release Q2', '2026-06-26', 200.00],
    )
