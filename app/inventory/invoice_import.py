"""Parse and import Payable Invoice Detail exports (Xero-style spreadsheet)."""
from __future__ import annotations

import logging
import re
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
from typing import Any

from django.utils import timezone as django_timezone
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

HEADER_ALIASES = {
    'invoice date': 'invoice_date',
    'source': 'source',
    'reference': 'reference',
    'item code': 'item_code',
    'description': 'description',
    'quantity': 'quantity',
    'qty': 'quantity',
    'units': 'quantity',
    'original currency': 'original_currency',
    'unit price (ex) (source)': 'unit_price_source',
    'gross (source)': 'gross_source',
    'unit price (ex) (myr)': 'unit_price_myr',
    'gross (myr)': 'gross_myr',
    'invoice total (myr)': 'invoice_total_myr',
}

# Standard Payable Invoice Detail column order (after description).
_STANDARD_COL_OFFSETS = {
    'quantity': 1,
    'original_currency': 2,
    'unit_price_source': 3,
    'gross_source': 4,
    'unit_price_myr': 5,
    'gross_myr': 6,
    'invoice_total_myr': 7,
}

SKIP_ROW_PREFIXES = (
    'payable invoice detail',
    'for the period',
    'branch is',
    'account contains',
    'status contains',
)


def _normalize_header(cell: Any) -> str:
    if cell is None:
        return ''
    text = str(cell).replace('\xa0', ' ').strip().lower()
    return ' '.join(text.split())


def _cell_at(row: tuple | list, idx: int | None) -> Any:
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _cell_str(row: tuple | list, idx: int | None) -> str:
    val = _cell_at(row, idx)
    if val is None:
        return ''
    return str(val).strip()


def _parse_decimal(val: Any) -> Decimal | None:
    if val is None or (isinstance(val, str) and not val.strip()):
        return None
    if isinstance(val, bool):
        return None
    if isinstance(val, Decimal):
        return val
    if isinstance(val, int):
        return Decimal(val)
    if isinstance(val, float):
        return Decimal(str(val))
    s = str(val).strip().replace('\xa0', '').replace(' ', '')
    if not s:
        return None
    # Single comma as decimal separator (e.g. 60,0000)
    if ',' in s and '.' not in s:
        s = s.replace(',', '.')
    # European-style decimals: 1.234,56
    elif ',' in s and '.' in s and s.rfind(',') > s.rfind('.'):
        s = s.replace('.', '').replace(',', '.')
    else:
        s = s.replace(',', '')
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _parse_quantity(val: Any) -> int:
    d = _parse_decimal(val)
    if d is None:
        return 0
    if d <= 0:
        return 0
    return int(d.quantize(Decimal('1'), rounding=ROUND_HALF_UP))


def _infer_quantity(
    qty: int,
    *,
    unit_price: Decimal | None,
    gross: Decimal | None,
    unit_price_myr: Decimal | None = None,
    gross_myr: Decimal | None = None,
) -> int:
    """Derive quantity from gross ÷ unit when the quantity cell is empty."""
    if qty > 0:
        return qty
    for unit, gross_val in (
        (unit_price, gross),
        (unit_price_myr, gross_myr),
    ):
        if unit and gross_val and unit > 0:
            inferred = (gross_val / unit).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
            if inferred > 0:
                return int(inferred)
    return 0


def _parse_date_cell(val: Any) -> date | None:
    if val is None or (isinstance(val, str) and not str(val).strip()):
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ('%d %b %Y', '%d %B %Y', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _map_headers_to_columns(headers: list[str]) -> dict[str, int]:
    col_map: dict[str, int] = {}
    for idx, h in enumerate(headers):
        if not h:
            continue
        if h in HEADER_ALIASES:
            col_map[HEADER_ALIASES[h]] = idx

    for idx, h in enumerate(headers):
        if not h:
            continue
        if 'quantity' not in col_map and (h in ('qty', 'units') or h == 'quantity' or h.startswith('qty')):
            col_map['quantity'] = idx
            continue
        if 'original_currency' not in col_map and (h.startswith('original currency') or h == 'currency'):
            col_map['original_currency'] = idx
            continue
        if 'unit_price_source' not in col_map and 'unit price' in h and 'source' in h:
            col_map['unit_price_source'] = idx
            continue
        if 'gross_source' not in col_map and 'gross' in h and 'source' in h:
            col_map['gross_source'] = idx
            continue
        if 'unit_price_myr' not in col_map and 'unit price' in h and 'myr' in h:
            col_map['unit_price_myr'] = idx
            continue
        if 'gross_myr' not in col_map and 'gross' in h and 'myr' in h and 'invoice total' not in h:
            col_map['gross_myr'] = idx
            continue
        if 'invoice_total_myr' not in col_map and 'invoice total' in h and 'myr' in h:
            col_map['invoice_total_myr'] = idx
    return col_map


def _fill_standard_column_gaps(col_map: dict[str, int]) -> None:
    """Infer missing columns from the standard Payable Invoice Detail layout."""
    desc_idx = col_map.get('description')
    if desc_idx is None:
        return
    for field, offset in _STANDARD_COL_OFFSETS.items():
        if field not in col_map:
            col_map[field] = desc_idx + offset


def _resolve_line_quantity(row: tuple | list, col_map: dict[str, int]) -> int:
    """Read quantity from the mapped column, or scan between description and currency."""
    qty_idx = col_map.get('quantity')
    qty = _parse_quantity(_cell_at(row, qty_idx))
    if qty > 0:
        return qty

    desc_idx = col_map.get('description')
    if desc_idx is None:
        return 0

    cur_idx = col_map.get('original_currency')
    if cur_idx is None:
        unit_idx = col_map.get('unit_price_source')
        cur_idx = unit_idx if unit_idx is not None else desc_idx + 2

    for idx in range(desc_idx + 1, cur_idx):
        if idx == qty_idx:
            continue
        candidate = _parse_quantity(_cell_at(row, idx))
        if candidate > 0:
            return candidate
    return 0


def _find_header_row(rows: list[tuple]) -> tuple[int | None, dict[str, int]]:
    best_i: int | None = None
    best_map: dict[str, int] = {}
    best_score = 0
    for i, row in enumerate(rows):
        headers = [_normalize_header(c) for c in row]
        col_map = _map_headers_to_columns(headers)
        if 'reference' not in col_map or 'description' not in col_map:
            continue
        _fill_standard_column_gaps(col_map)
        score = len(col_map)
        if score > best_score:
            best_score = score
            best_i = i
            best_map = col_map
    return best_i, best_map


def _is_metadata_or_blank_row(row: tuple) -> bool:
    if not row or all(c is None or str(c).strip() == '' for c in row):
        return True
    first = _cell_str(row, 0).lower()
    if not first:
        return True
    return any(first.startswith(p) for p in SKIP_ROW_PREFIXES)


def _is_supplier_row(row: tuple, col_map: dict[str, int]) -> bool:
    if _is_metadata_or_blank_row(row):
        return False
    name = _cell_str(row, 0)
    if not name:
        return False
    if _parse_date_cell(_cell_at(row, col_map.get('invoice_date'))) if 'invoice_date' in col_map else _parse_date_cell(name):
        return False
    ref = _cell_str(row, col_map.get('reference'))
    desc = _cell_str(row, col_map.get('description'))
    if ref or desc:
        return False
    lowered = name.lower()
    if lowered in HEADER_ALIASES or lowered in {k for k in HEADER_ALIASES.values()}:
        return False
    return True


def _is_line_row(row: tuple, col_map: dict[str, int]) -> bool:
    ref = _cell_str(row, col_map.get('reference'))
    desc = _cell_str(row, col_map.get('description'))
    return bool(ref and desc)


def _supplier_key(name: str) -> str:
    return ' '.join(name.strip().lower().split())


def _empty_invoice() -> dict:
    return {
        'reference': '',
        'invoice_date': None,
        'invoice_total_myr': None,
        'lines': [],
    }


def _code_from_supplier_name(name: str) -> str:
    """First four alphabetic characters of the supplier name, uppercased."""
    letters = re.findall(r'[A-Za-z]', name or '')
    return ''.join(letters[:4]).upper()


def suggest_supplier_code(
    file_supplier_name: str,
    supplier_model=None,
    *,
    invoice_refs: list[str] | None = None,
) -> str:
    """Derive supplier code from the first 4 letters of the file group title."""
    code = _code_from_supplier_name(file_supplier_name)
    if not code:
        return ''
    if not supplier_model:
        return code

    base = code
    n = 2
    while supplier_model.objects.filter(code__iexact=code).exists():
        suffix = str(n)
        code = (base[: max(1, 50 - len(suffix))] + suffix)[:50]
        n += 1
    return code


def _read_xlsx_rows(file) -> list[tuple]:
    if hasattr(file, 'read'):
        content = file.read()
        if hasattr(file, 'seek'):
            file.seek(0)
    else:
        content = file
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    raw_rows: list[tuple] = []
    max_len = 0
    for row in ws.iter_rows(values_only=True):
        row_tuple = tuple(row)
        max_len = max(max_len, len(row_tuple))
        raw_rows.append(row_tuple)
    wb.close()
    if max_len == 0:
        return []
    return [
        row + (None,) * (max_len - len(row))
        for row in raw_rows
    ]


def parse_payable_invoice_detail_file(file) -> tuple[dict | None, str | None]:
    """
    Parse a Payable Invoice Detail .xlsx export.
    Returns ({ suppliers: [...], summary: {...} }, None) or (None, error_message).
    """
    name = getattr(file, 'name', '') or ''
    if not name.lower().endswith(('.xlsx', '.xls')):
        return None, 'Please upload an Excel file (.xlsx) exported as Payable Invoice Detail.'

    try:
        rows = _read_xlsx_rows(file)
    except Exception as exc:
        return None, f'Could not read Excel file: {exc}'

    header_idx, col_map = _find_header_row(rows)
    if header_idx is None:
        return None, (
            'Could not find the invoice table header row. '
            'Expected columns such as Reference and Description.'
        )

    current_supplier_key: str | None = None
    supplier_buckets: dict[str, dict] = {}

    for row in rows[header_idx + 1:]:
        if _is_metadata_or_blank_row(row):
            continue
        if _is_supplier_row(row, col_map):
            supplier_name = _cell_str(row, 0)
            bucket_key = _supplier_key(supplier_name)
            if bucket_key not in supplier_buckets:
                supplier_buckets[bucket_key] = {
                    'file_supplier_name': supplier_name,
                    'invoices': defaultdict(_empty_invoice),
                }
            current_supplier_key = bucket_key
            continue
        if not _is_line_row(row, col_map):
            continue
        if not current_supplier_key:
            return None, (
                f'Line item "{_cell_str(row, col_map.get("description"))}" '
                'appears before any supplier group title.'
            )

        invoices_map = supplier_buckets[current_supplier_key]['invoices']
        ref = _cell_str(row, col_map.get('reference'))
        inv = invoices_map[ref]
        inv['reference'] = ref
        inv_date = _parse_date_cell(_cell_at(row, col_map.get('invoice_date')))
        if inv_date and not inv['invoice_date']:
            inv['invoice_date'] = inv_date.isoformat()
        total_myr = _parse_decimal(_cell_at(row, col_map.get('invoice_total_myr')))
        if total_myr is not None:
            inv['invoice_total_myr'] = float(total_myr)

        unit_source = _parse_decimal(_cell_at(row, col_map.get('unit_price_source')))
        gross_source = _parse_decimal(_cell_at(row, col_map.get('gross_source')))
        unit_myr = _parse_decimal(_cell_at(row, col_map.get('unit_price_myr')))
        gross_myr = _parse_decimal(_cell_at(row, col_map.get('gross_myr')))
        qty = _resolve_line_quantity(row, col_map)
        qty = _infer_quantity(
            qty,
            unit_price=unit_source,
            gross=gross_source,
            unit_price_myr=unit_myr,
            gross_myr=gross_myr,
        )
        if unit_myr is None and gross_myr is not None and qty > 0:
            unit_myr = (gross_myr / Decimal(qty)).quantize(Decimal('0.01'))
        if unit_source is None and gross_source is not None and qty > 0:
            unit_source = (gross_source / Decimal(qty)).quantize(Decimal('0.01'))

        inv['lines'].append({
            'description': _cell_str(row, col_map.get('description')),
            'item_code': _cell_str(row, col_map.get('item_code')),
            'quantity': qty,
            'unit_price_myr': float(unit_myr) if unit_myr is not None else None,
            'gross_myr': float(gross_myr) if gross_myr is not None else None,
            'unit_price_source': float(unit_source) if unit_source is not None else None,
            'gross_source': float(gross_source) if gross_source is not None else None,
            'original_currency': _cell_str(row, col_map.get('original_currency')),
        })

    if not supplier_buckets:
        return None, 'No invoice line items found in file.'

    suppliers_out = []
    total_invoices = 0
    total_lines = 0
    for bucket_key, bucket in supplier_buckets.items():
        invoices_map = bucket['invoices']
        file_supplier_name = bucket['file_supplier_name']
        invoices_list = []
        for inv_data in invoices_map.values():
            if not inv_data['lines']:
                continue
            invoices_list.append(inv_data)
        if not invoices_list:
            continue
        line_count = sum(len(inv['lines']) for inv in invoices_list)
        total_invoices += len(invoices_list)
        total_lines += line_count
        suppliers_out.append({
            'key': bucket_key,
            'file_supplier_name': file_supplier_name,
            'invoice_count': len(invoices_list),
            'line_count': line_count,
            'invoices': invoices_list,
        })

    if not suppliers_out:
        return None, 'No invoice line items found in file.'

    return {
        'suppliers': suppliers_out,
        'summary': {
            'supplier_count': len(suppliers_out),
            'invoice_count': total_invoices,
            'line_count': total_lines,
        },
    }, None


def suggest_supplier_match(file_supplier_name: str, supplier_model) -> dict:
    """Suggest a system supplier for a file group title."""
    exact = supplier_model.objects.filter(name__iexact=file_supplier_name).first()
    if exact:
        return {'supplier_id': exact.pk, 'supplier_name': exact.name, 'match_type': 'exact'}
    first_word = (file_supplier_name or '').strip().split()[0] if file_supplier_name else ''
    if len(first_word) >= 3:
        partial = supplier_model.objects.filter(name__icontains=first_word).order_by('name').first()
        if partial:
            return {'supplier_id': partial.pk, 'supplier_name': partial.name, 'match_type': 'fuzzy'}
    return {'supplier_id': None, 'supplier_name': None, 'match_type': 'none'}


def _normalize_import_action(action: str | None) -> str:
    return (action or 'map').strip().lower()


def _safe_invoice_id(reference: str) -> str:
    ref = (reference or '').strip()
    if not ref:
        raise ValueError('Invoice reference is required.')
    if len(ref) <= 20:
        return ref
    raise ValueError(
        f'Invoice reference "{ref}" is too long ({len(ref)} characters). '
        'Maximum length is 20 characters.'
    )


def _resolve_or_create_supplier(sup_block: dict, supplier_model, *, import_cache: dict | None = None):
    """Resolve supplier for one import group; reuse by code within the same import batch."""
    import_cache = import_cache if import_cache is not None else {}
    action = _normalize_import_action(sup_block.get('action'))
    if action == 'ignore':
        return None, action

    if action == 'create':
        new_name = (sup_block.get('new_supplier_name') or sup_block.get('file_supplier_name') or '').strip()
        if not new_name:
            raise ValueError(f'Supplier name required for "{sup_block.get("file_supplier_name")}".')
        new_code = (sup_block.get('new_supplier_code') or '').strip() or None
        code_key = new_code.upper() if new_code else None

        if code_key:
            if code_key in import_cache:
                return import_cache[code_key], action
            existing_by_code = supplier_model.objects.filter(code__iexact=new_code).first()
            if existing_by_code:
                import_cache[code_key] = existing_by_code
                return existing_by_code, action

        supplier, created = supplier_model.objects.get_or_create(name=new_name)

        if code_key:
            conflict = supplier_model.objects.filter(code__iexact=new_code).exclude(pk=supplier.pk).first()
            if conflict:
                import_cache[code_key] = conflict
                return conflict, action
            if created or not supplier.code:
                supplier.code = new_code
                supplier.save(update_fields=['code'])
            import_cache[code_key] = supplier

        return supplier, action

    sid = sup_block.get('supplier_id')
    if sid in (None, ''):
        raise ValueError(
            f'Choose a system supplier for "{sup_block.get("file_supplier_name")}" or mark as ignore.'
        )
    try:
        sid_int = int(sid)
    except (TypeError, ValueError):
        raise ValueError(f'Invalid supplier id for "{sup_block.get("file_supplier_name")}".')
    supplier = supplier_model.objects.filter(pk=sid_int).first()
    if not supplier:
        raise ValueError(f'Supplier id {sid_int} not found.')
    return supplier, action


def _matrix_tier_snapshot(unit_myr, unit_source=None) -> list[dict]:
    tier: dict = {
        'min_quantity': 1,
        'max_quantity': None,
        'unit_price': str(Decimal(str(unit_myr)).quantize(Decimal('0.01'))),
    }
    if unit_source is not None:
        tier['unit_price_source'] = str(Decimal(str(unit_source)).quantize(Decimal('0.0001')))
    return [tier]


def _refresh_matrix_entry_from_latest_record(entry) -> None:
    """Set live tiers on a matrix entry from its newest effective-dated upload record."""
    from inventory.models import SupplierPriceMatrixTier

    record = entry.upload_records.order_by('-effective_date', '-uploaded_at').first()
    if not record:
        return
    entry.tiers.all().delete()
    for tier in record.tiers or []:
        SupplierPriceMatrixTier.objects.create(
            entry=entry,
            min_quantity=int(tier.get('min_quantity', 1)),
            max_quantity=tier.get('max_quantity'),
            unit_price=Decimal(str(tier['unit_price'])),
        )
    entry.price_currency = record.price_currency
    entry.conversion_rate = record.conversion_rate
    entry.effective_date = record.effective_date
    entry.source_filename = record.source_filename
    entry.save(update_fields=[
        'price_currency', 'conversion_rate', 'effective_date', 'source_filename', 'updated_at',
    ])


def _upsert_supplier_price_matrix_from_line(
    supplier,
    product,
    line: dict,
    *,
    invoice_date: date | None,
    invoice_reference: str,
    source_filename: str = '',
) -> None:
    """Append invoice-priced snapshot to supplier matrix upload history."""
    from inventory.models import SupplierPriceMatrixEntry, SupplierPriceMatrixUploadRecord
    from inventory.supplier_pricing import sync_saved_base_costs_for_products

    desc = (line.get('description') or '').strip()
    if not desc:
        return
    unit_myr = line.get('unit_price_myr')
    if unit_myr is None:
        unit_myr = line.get('unit_price_source')
    if unit_myr is None:
        return

    currency = (line.get('original_currency') or 'MYR').strip().upper()[:3] or 'MYR'
    if currency not in ('MYR', 'USD', 'EUR'):
        currency = 'MYR'

    conversion_rate = None
    unit_source = line.get('unit_price_source')
    unit_myr_dec = Decimal(str(unit_myr))
    if unit_source and currency != 'MYR' and unit_myr_dec > 0:
        try:
            conversion_rate = (unit_myr_dec / Decimal(str(unit_source))).quantize(Decimal('0.0001'))
        except (InvalidOperation, ZeroDivisionError):
            conversion_rate = None

    entry, _created = SupplierPriceMatrixEntry.objects.update_or_create(
        supplier=supplier,
        line_medication=desc[:255],
        strength='',
        size='',
        defaults={
            'form': '',
            'product': product,
        },
    )
    if product and entry.product_id != product.pk:
        entry.product = product
        entry.save(update_fields=['product', 'updated_at'])

    effective = invoice_date or django_timezone.now().date()
    invoice_label = (invoice_reference or '').strip()
    record_source = source_filename or 'payable invoice import'
    if invoice_label:
        record_source = f'{record_source} · {invoice_label}'

    snapshot = _matrix_tier_snapshot(unit_myr_dec, unit_source=unit_source)
    existing_record = entry.upload_records.filter(
        effective_date=effective,
        source_filename=record_source,
    ).first()
    if not existing_record:
        SupplierPriceMatrixUploadRecord.objects.create(
            entry=entry,
            effective_date=effective,
            source_filename=record_source,
            price_currency=currency,
            conversion_rate=conversion_rate,
            tiers=snapshot,
        )

    _refresh_matrix_entry_from_latest_record(entry)

    if product:
        sync_saved_base_costs_for_products([product.pk])


def confirm_payable_invoice_import(payload: dict, *, product_model, supplier_model, invoice_model, invoice_item_model) -> dict:
    """
    Create/update standalone invoices from preview payload and supplier mappings.
    Returns counts for toast messaging.
    """
    from sales.models import Invoice

    products_created = 0
    products_matched = 0
    invoices_created = 0
    invoices_updated = 0
    invoices_skipped = 0
    suppliers_ignored = 0
    matrix_rows_updated = 0
    supplier_import_cache: dict[str, object] = {}

    for sup_block in payload.get('suppliers') or []:
        supplier, action = _resolve_or_create_supplier(
            sup_block, supplier_model, import_cache=supplier_import_cache,
        )
        if action == 'ignore' or supplier is None:
            suppliers_ignored += 1
            continue

        for inv_data in sup_block.get('invoices') or []:
            reference = (inv_data.get('reference') or '').strip()
            if not reference:
                continue

            invoice_id = _safe_invoice_id(reference)

            existing = invoice_model.objects.filter(invoice_id=invoice_id).select_related('quotation').first()
            if existing and existing.quotation_id is not None:
                invoices_skipped += 1
                continue

            inv_date_raw = inv_data.get('invoice_date')
            if inv_date_raw:
                inv_date = datetime.strptime(inv_date_raw[:10], '%Y-%m-%d').date()
            else:
                inv_date = django_timezone.now().date()

            lines = inv_data.get('lines') or []
            subtotal = Decimal('0')
            for line in lines:
                gross = line.get('gross_myr')
                if gross is not None:
                    subtotal += Decimal(str(gross))
                else:
                    qty = int(line.get('quantity') or 0)
                    unit = line.get('unit_price_myr')
                    if unit is not None and qty:
                        subtotal += Decimal(str(unit)) * qty

            invoice_total = inv_data.get('invoice_total_myr')
            transport = Decimal('0')
            if invoice_total is not None:
                transport = max(Decimal('0'), Decimal(str(invoice_total)) - subtotal)

            if existing:
                invoice = existing
                invoice.supplier = supplier
                invoice.date_issued = inv_date
                invoice.transportation_cost = transport
                invoice.status = Invoice.InvoiceStatus.PAID
                invoice.save()
                invoices_updated += 1
                invoice.items.all().delete()
            else:
                invoice = invoice_model.objects.create(
                    invoice_id=invoice_id,
                    quotation=None,
                    supplier=supplier,
                    date_issued=inv_date,
                    transportation_cost=transport,
                    status=Invoice.InvoiceStatus.PAID,
                    notes=f'Imported from {payload.get("source_filename") or "spreadsheet"}',
                )
                invoices_created += 1

            items_to_create = []
            for line in lines:
                desc = (line.get('description') or '').strip()
                if not desc:
                    continue
                item_code = (line.get('item_code') or '').strip()
                product = None
                if item_code:
                    product = product_model.objects.filter(sku__iexact=item_code).first()
                if not product:
                    product = product_model.objects.filter(name__iexact=desc).first()
                if not product:
                    name = desc[:200]
                    product, created = product_model.objects.get_or_create(
                        name=name,
                        defaults={'description': 'Auto-imported from payable invoice'},
                    )
                    if created:
                        products_created += 1
                    else:
                        products_matched += 1
                else:
                    products_matched += 1

                if supplier and not product.suppliers.filter(pk=supplier.pk).exists():
                    product.suppliers.add(supplier)

                qty = _parse_quantity(line.get('quantity'))
                unit_price = line.get('unit_price_myr')
                if unit_price is None:
                    unit_price = line.get('unit_price_source')
                if unit_price is None:
                    continue
                original_currency = (line.get('original_currency') or '').strip().upper()[:3]
                unit_price_source = line.get('unit_price_source')
                gross_source = line.get('gross_source')
                item_kwargs = {
                    'invoice': invoice,
                    'product': product,
                    'description': desc,
                    'quantity': max(qty, 0),
                    'unit_price': Decimal(str(unit_price)),
                }
                if hasattr(invoice_item_model, 'original_currency'):
                    item_kwargs['original_currency'] = original_currency
                if hasattr(invoice_item_model, 'unit_price_source') and unit_price_source is not None:
                    item_kwargs['unit_price_source'] = Decimal(str(unit_price_source))
                if hasattr(invoice_item_model, 'gross_source') and gross_source is not None:
                    item_kwargs['gross_source'] = Decimal(str(gross_source))
                items_to_create.append(invoice_item_model(**item_kwargs))
                try:
                    _upsert_supplier_price_matrix_from_line(
                        supplier,
                        product,
                        line,
                        invoice_date=inv_date,
                        invoice_reference=reference,
                        source_filename=payload.get('source_filename') or '',
                    )
                    matrix_rows_updated += 1
                except Exception as exc:
                    logger.warning('Supplier price matrix update failed for %s: %s', desc, exc)
            if items_to_create:
                invoice_item_model.objects.bulk_create(items_to_create)

    return {
        'products_created': products_created,
        'products_matched': products_matched,
        'invoices_created': invoices_created,
        'invoices_updated': invoices_updated,
        'invoices_skipped': invoices_skipped,
        'suppliers_ignored': suppliers_ignored,
        'matrix_rows_updated': matrix_rows_updated,
    }
